import os
import re
import shutil
import logging
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.utils import column_index_from_string
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from typing import List
from fastapi.middleware.cors import CORSMiddleware

# Logging Configuration
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# Configuration Dictionary
CONFIG = {
    "month_column_ranges": {
        "November": ("CC", "DF"),
        "December": ("DG", "EK"),
        "January": ("EL", "FP"),
        "February": ("FQ", "GR"),
        "March": ("GS", "HW"),
        "April": ("HX", "JA"),
        "May": ("JB", "KF")
    },
    "new_comment_columns": {
        "INVALIDITY": 5,
        "H_ALM": 7,
        "FLAME": 10,
        "HH_ALM": 6,
        "BEAM": 8,
        "OTHERS": 9,
    },
    "filter_keywords": ["GDB", "NFD", "NSD", "GDIR", "NMAC"],
    "columns_to_extract": ["TAG", "DESCRIPTION", "ALARM DESC2", "ACT/UNACK"],
}

# FastAPI App
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow requests from anywhere (for testing)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure temporary storage
TEMP_DIR = "/tmp" if "RENDER" in os.environ else os.path.join(os.getcwd(), "uploads")
os.makedirs(TEMP_DIR, exist_ok=True)


def extract_date_and_sheet(file_path):
    try:
        match = re.search(r"(\d{4}-\d{2}-\d{2})", file_path)
        if not match:
            raise ValueError(f"No valid date found in '{file_path}'")
        date_str = match.group(1)
        formatted_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        target_sheet = f"All_{formatted_date.replace('/', '-')}"

        return formatted_date, target_sheet
    except Exception as e:
        logging.error(f"Error extracting date and sheet: {e}")
        raise HTTPException(status_code=400, detail="Invalid file name format")


def load_and_filter_data(file_path, target_sheet):
    try:
        engine = "xlrd" if file_path.endswith(".xls") else "openpyxl"
        excel_data = pd.ExcelFile(file_path, engine=engine)
        if target_sheet not in excel_data.sheet_names:
            raise ValueError(f"Sheet '{target_sheet}' not found in '{file_path}'")
        data = pd.read_excel(file_path, sheet_name=target_sheet, engine=engine)
        for col in CONFIG["columns_to_extract"]:
            if col not in data.columns:
                raise ValueError(f"Column '{col}' not found in the data from '{file_path}'")
        filtered_data = data[data["TAG"].str.contains("|".join(CONFIG["filter_keywords"]), na=False)]
        return filtered_data[CONFIG["columns_to_extract"]]
    except ValueError as ve:
        logging.error(f"Data extraction error: {ve}")
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logging.error(f"Error loading data from '{file_path}': {e}")
        raise HTTPException(status_code=500, detail="Data extraction failed")


def convert_xls_to_xlsx(xls_path):
    try:
        df = pd.read_excel(xls_path, engine="xlrd")
        xlsx_path = xls_path + "x"  # e.g., "tracker.xls" → "tracker.xlsx"
        df.to_excel(xlsx_path, index=False, engine="openpyxl")
        logging.info(f"Converted '{xls_path}' to '{xlsx_path}'")
        return xlsx_path
    except Exception as e:
        logging.error(f"Failed to convert {xls_path} to xlsx: {e}")
        raise HTTPException(status_code=500, detail="Failed to convert .xls tracker file")


def update_tracker(tracker_path, extracted_dfs, is_latest): 
    try:
        workbook = openpyxl.load_workbook(tracker_path)
        sheet = workbook.active

        updated_rows = set()
        last_tag_row = sheet.max_row

        # Step 1: Reset all YES/NO columns to "NO" before updating
        logging.info("Resetting all YES/NO columns to NO before updating...")
        for i in range(12, sheet.max_row + 1):  # Start from row 12
            for col in CONFIG["new_comment_columns"].values():
                sheet.cell(row=i, column=col).value = "NO"  # ✅ Ensures a clean reset

        for extracted_df, formatted_date in extracted_dfs:
            parsed_date = datetime.strptime(formatted_date, "%d/%m/%Y")
            month = parsed_date.strftime("%B")
            day = parsed_date.day

            if month not in CONFIG["month_column_ranges"]:
                raise ValueError(f"Invalid month: {month}.")
            
            start_col_letter, end_col_letter = CONFIG["month_column_ranges"][month]
            start_col = column_index_from_string(start_col_letter)
            end_col = column_index_from_string(end_col_letter)

            day_column = None
            for col in range(start_col, end_col + 1):
                if sheet.cell(row=11, column=col).value == day:
                    day_column = col
                    break
            if not day_column:
                raise HTTPException(status_code=400, detail=f"Day {day} not found in {month}. Check tracker format.")

            # Step 2: Process extracted rows
            for _, row in extracted_df.iterrows():
                tag = row["TAG"]
                location = row["DESCRIPTION"]
                act_unack = row["ACT/UNACK"]
                alarm_desc = row["ALARM DESC2"]

                new_comment = {
                    "Invalidity State": "INVALIDITY",
                    "High Gas Level Detection": "H_ALM",
                    "Flame Detection": "FLAME",
                    "High High Gas level Detection": "HH_ALM",
                    "SMOKE Detection": "BEAM",
                }.get(alarm_desc, "OTHERS")

                formatted_comment = f"{formatted_date} {new_comment}"

                # Check if tag already exists
                tag_lookup = {sheet.cell(row=i, column=2).value: i for i in range(12, sheet.max_row + 1)}
                tag_row = tag_lookup.get(tag, None)

                if tag_row:
                    # Update existing tag row
                    sheet.cell(row=tag_row, column=day_column).value = act_unack
                    comment_col_index = sheet.max_column - 2
                    existing_comment = sheet.cell(row=tag_row, column=comment_col_index).value or ""
                    sheet.cell(row=tag_row, column=comment_col_index).value = f"{existing_comment}\n{formatted_comment}".strip()

                    # Step 3: Assign "YES" after resetting
                    if is_latest:
                        col_number = CONFIG["new_comment_columns"].get(new_comment)
                        if col_number:
                            sheet.cell(row=tag_row, column=col_number).value = "YES"
                            
                else:
                    # Append new row for new tags
                    last_tag_row += 1
                    sheet.cell(row=last_tag_row, column=2).value = tag
                    sheet.cell(row=last_tag_row, column=3).value = location
                    sheet.cell(row=last_tag_row, column=day_column).value = act_unack
                    sheet.cell(row=last_tag_row, column=sheet.max_column - 2).value = formatted_comment

                    # Step 4: Set all YES/NO columns to "NO" for new tags
                    for col in CONFIG["new_comment_columns"].values():
                        sheet.cell(row=last_tag_row, column=col).value = "NO"

                    # Step 5: Assign "YES" for new tags if required
                    if is_latest:
                        col_number = CONFIG["new_comment_columns"].get(new_comment)
                        if col_number:
                            sheet.cell(row=last_tag_row, column=col_number).value = "YES"

        temp_output_path = os.path.join(TEMP_DIR, "updated_tracker.xlsx")
        workbook.save(temp_output_path)
        return temp_output_path
    except ValueError as ve:
        logging.error(f"Tracker update error: {ve}")
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logging.error(f"Error updating tracker: {e}")
        raise HTTPException(status_code=500, detail="Tracker update failed")


@app.post("/process_backlog")
async def process_backlog(input_files: List[UploadFile] = File(...), tracker_file: UploadFile = File(...)):
    try:
        extracted_dfs = []

        # Save tracker file
        tracker_path = os.path.join(TEMP_DIR, tracker_file.filename)
        with open(tracker_path, "wb") as buffer:
            shutil.copyfileobj(tracker_file.file, buffer)

        # Convert tracker if it’s .xls
        if tracker_path.endswith(".xls"):
            converted_path = convert_xls_to_xlsx(tracker_path)
            if not converted_path:
                raise HTTPException(status_code=500, detail="Failed to convert .xls tracker file.")
            tracker_path = converted_path

        # Loop through and handle each input file
        for input_file in input_files:
            input_path = os.path.join(TEMP_DIR, input_file.filename)
            with open(input_path, "wb") as buffer:
                shutil.copyfileobj(input_file.file, buffer)

            formatted_date, target_sheet = extract_date_and_sheet(input_path)
            if not formatted_date:
                raise HTTPException(status_code=400, detail=f"Invalid file name format: {input_file.filename}")

            extracted_df = load_and_filter_data(input_path, target_sheet)
            if extracted_df is None:
                raise HTTPException(status_code=500, detail=f"Data extraction failed for {input_file.filename}")

            extracted_dfs.append((extracted_df, formatted_date))

        # Update tracker
        updated_tracker_path = update_tracker(tracker_path, extracted_dfs, is_latest=False)
        if not updated_tracker_path:
            raise HTTPException(status_code=500, detail="Tracker update failed.")

        return FileResponse(updated_tracker_path, filename="updated_backlog_tracker.xlsx")
    except Exception as e:
        logging.error(f"Error processing backlog: {e}")
        raise HTTPException(status_code=500, detail="Error processing backlog")


@app.post("/process_latest")
async def process_latest(input_file: UploadFile = File(...), tracker_file: UploadFile = File(...)):
    try:
        extracted_dfs = []

        # Save tracker file
        tracker_path = os.path.join(TEMP_DIR, tracker_file.filename)
        with open(tracker_path, "wb") as buffer:
            shutil.copyfileobj(tracker_file.file, buffer)

        # Convert tracker if it’s .xls
        if tracker_path.endswith(".xls"):
            converted_path = convert_xls_to_xlsx(tracker_path)
            if not converted_path:
                raise HTTPException(status_code=500, detail="Failed to convert .xls tracker file.")
            tracker_path = converted_path

        # Save input file
        input_path = os.path.join(TEMP_DIR, input_file.filename)
        with open(input_path, "wb") as buffer:
            shutil.copyfileobj(input_file.file, buffer)

        formatted_date, target_sheet = extract_date_and_sheet(input_path)
        if not formatted_date:
            raise HTTPException(status_code=400, detail=f"Invalid file name format: {input_file.filename}")

        extracted_df = load_and_filter_data(input_path, target_sheet)
        if extracted_df is None:
            raise HTTPException(status_code=500, detail=f"Data extraction failed for {input_file.filename}")

        extracted_dfs.append((extracted_df, formatted_date))

        updated_tracker_path = update_tracker(tracker_path, extracted_dfs, is_latest=True)
        if not updated_tracker_path:
            raise HTTPException(status_code=500, detail="Tracker update failed.")

        return FileResponse(updated_tracker_path, filename="updated_latest_tracker.xlsx")
    except Exception as e:
        logging.error(f"Error processing latest: {e}")
        raise HTTPException(status_code=500, detail="Error processing latest tracker")
